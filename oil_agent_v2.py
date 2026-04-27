#!/usr/bin/env python3
"""
Oil Price Agent v2 -- Deterministic, no LLM.
Fetches live WTI/Brent data, runs a log-normal + bootstrap ensemble,
generates charts and a .txt report. Called by app.py via run().
"""

import os, io, json, uuid, base64, datetime, warnings
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from dateutil import parser as dateparser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import data_fetcher as _df

warnings.filterwarnings('ignore')

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

HORIZONS     = ["1M", "3M", "6M", "9M", "12M"]
HORIZON_DAYS = {"1M": 21, "3M": 63, "6M": 126, "9M": 189, "12M": 252}

GAS_PRICE_BINS = [
    "<3.30", "3.30-3.40", "3.40-3.50", "3.50-3.60",
    "3.60-3.70", "3.70-3.90", "3.90-4.10", "4.10-4.30",
    "4.30-4.60", "4.60-5.00", "5.00-5.50", ">5.50",
]
GAS_BIN_EDGES = [-np.inf, 3.30, 3.40, 3.50, 3.60,
                  3.70, 3.90, 4.10, 4.30, 4.60, 5.00, 5.50, np.inf]

WTI_TO_GAS_MARKUP = 1.15  # approximate retail gas markup over WTI/42


# -- PROBABILITY MODELS --------------------------------------------------------

def lognormal_probs(current, returns, horizon_days, bin_edges):
    """
    Log-normal probability distribution for commodity prices.
    mu=0 (martingale/risk-neutral): oil is a commodity, not an equity.
    We do NOT extrapolate historical trend into the future.
    Only volatility (sigma) drives the distribution width.
    Sigma is annualised then scaled: sig_daily = sig_annual / sqrt(252).
    """
    from scipy.stats import norm
    r   = np.array(returns) if len(returns) > 5 else np.random.normal(0, 0.012, 60)
    # Use only vol — no drift. Oil is mean-reverting, not trending.
    sig_daily = float(np.std(r, ddof=1)) or 0.012
    # Cap annualised vol at 80% to prevent extreme tail widening
    sig_annual_cap = 0.80
    sig_daily = min(sig_daily, sig_annual_cap / np.sqrt(252))
    # mu=0: forward price equals spot (commodity martingale assumption)
    log_mu  = np.log(current) - 0.5 * sig_daily**2 * horizon_days
    log_sig = sig_daily * np.sqrt(horizon_days)
    probs = []
    for i in range(len(bin_edges) - 1):
        lo, hi = bin_edges[i], bin_edges[i+1]
        p_lo = norm.cdf(np.log(max(lo, 1e-6)), loc=log_mu, scale=log_sig) if lo > -np.inf else 0.0
        p_hi = norm.cdf(np.log(hi),            loc=log_mu, scale=log_sig) if hi <  np.inf else 1.0
        probs.append(max(0.0, p_hi - p_lo))
    t = sum(probs) or 1.0
    return [p/t for p in probs]


def bootstrap_probs(current, returns, horizon_days, bin_edges, n=3000):
    """
    Historical simulation bootstrap.
    Returns are DEMEANED before sampling to remove trend bias.
    We keep only the volatility structure, not any historical directional drift.
    """
    r = np.array(returns) if len(returns) > 5 else np.random.normal(0, 0.012, 60)
    r = r - np.mean(r)  # demean: remove historical trend, keep only vol structure
    sims  = np.exp(np.sum(np.random.choice(r, size=(n, horizon_days), replace=True), axis=1))
    final = current * sims
    probs = []
    for i in range(len(bin_edges) - 1):
        lo, hi = bin_edges[i], bin_edges[i+1]
        mask = (True if lo == -np.inf else final >= lo) & (True if hi == np.inf else final < hi)
        probs.append(float(np.sum(mask)) / n)
    t = sum(probs) or 1.0
    return [p/t for p in probs]


def ensemble(p1, p2, w1=0.55, w2=0.45):
    combined = [w1*a + w2*b for a, b in zip(p1, p2)]
    t = sum(combined) or 1.0
    return [p/t for p in combined]


# -- FORECAST ------------------------------------------------------------------

def compute_forecast(wti, returns):
    """
    5-day log-normal forecast with mu=0.
    For oil/commodities, the best short-term forecast of price is current price
    (martingale). The CI widens symmetrically around today's price.
    """
    r   = np.array(returns) if len(returns) > 5 else np.random.normal(0, 0.012, 60)
    sig = min(float(np.std(r, ddof=1)) or 0.012, 0.80 / np.sqrt(252))
    h   = 5  # trading days
    # mu=0: no directional bias
    log_mu  = np.log(wti) - 0.5*sig**2 * h
    log_sig = sig * np.sqrt(h)
    mid  = round(float(np.exp(log_mu)), 2)
    low  = round(float(np.exp(log_mu - 1.645*log_sig)), 2)
    high = round(float(np.exp(log_mu + 1.645*log_sig)), 2)
    direction = "BULLISH" if mid > wti*1.005 else "BEARISH" if mid < wti*0.995 else "NEUTRAL"
    return {
        "current_wti":       round(wti, 2),
        "forecast_low":      low,
        "forecast_high":     high,
        "forecast_midpoint": mid,
        "direction":         direction,
        "annualised_vol":    round(sig * np.sqrt(252) * 100, 1),
        "model":             "55% log-normal + 45% bootstrap (5-day, 90% CI)",
    }


def build_wti_prob_table(wti, returns):
    edges  = list(range(40, 210, 10))
    b_edges  = [-np.inf] + edges + [np.inf]
    b_labels = ["<$40"] + ["${}-{}".format(edges[i], edges[i+1])
                            for i in range(len(edges)-1)] + [">${}".format(edges[-1])]
    table = {}
    for h in HORIZONS:
        try:
            p1 = lognormal_probs(wti, returns, HORIZON_DAYS[h], b_edges)
            p2 = bootstrap_probs(wti, returns, HORIZON_DAYS[h], b_edges)
            table[h] = list(zip(b_labels, ensemble(p1, p2)))
        except Exception:
            table[h] = [(l, 1/len(b_labels)) for l in b_labels]
    return table


# -- RULE-BASED SUMMARY --------------------------------------------------------

def market_summary(wti, brent, history):
    prices = [r["price"] for r in history]
    if len(prices) < 20:
        return "Insufficient history for summary."
    ma7  = np.mean(prices[-7:])
    ma20 = np.mean(prices[-20:])
    vol  = float(np.std(np.diff(np.log(prices[-20:])), ddof=1) * np.sqrt(252) * 100)
    chg1w = (prices[-1]-prices[-5]) /prices[-5]  *100 if len(prices)>=5  else 0
    chg1m = (prices[-1]-prices[-20])/prices[-20] *100 if len(prices)>=20 else 0
    spread = round(brent - wti, 2) if brent else 0
    trend  = "UPTREND"   if ma7 > ma20 else "DOWNTREND"
    vol_r  = "HIGH"      if vol > 35   else "ELEVATED" if vol > 22 else "NORMAL"
    mom    = "BULLISH"   if chg1w > 1  else "BEARISH"  if chg1w < -1 else "NEUTRAL"
    lines = [
        "=== OIL MARKET SUMMARY -- {} ===".format(datetime.date.today()), "",
        "PRICE SNAPSHOT",
        "  WTI Crude  : ${:.2f}/bbl".format(wti),
        "  Brent Crude: ${:.2f}/bbl".format(brent),
        "  Brent-WTI  : ${:+.2f}/bbl".format(spread), "",
        "TREND & MOMENTUM",
        "  Regime     : {}  (7d MA ${:.2f} vs 20d MA ${:.2f})".format(trend, ma7, ma20),
        "  1-week chg : {:+.1f}%".format(chg1w),
        "  1-month chg: {:+.1f}%".format(chg1m),
        "  Momentum   : {}".format(mom), "",
        "VOLATILITY",
        "  Annualised : {:.1f}%  [{}]".format(vol, vol_r), "",
        "MODEL",
        "  Ensemble: 55% log-normal + 45% historical bootstrap",
        "  Probabilities are model outputs. Not financial advice.",
    ]
    return "\n".join(lines)


# -- CHARTS --------------------------------------------------------------------

def _b64(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    buf.seek(0); plt.close(fig)
    return base64.b64encode(buf.read()).decode()


def chart_price(history, wti, f_low, f_high, f_mid):
    dates, prices = [], []
    for row in history:
        try:
            dates.append(dateparser.parse(str(row["date"])).date())
            prices.append(float(row["price"]))
        except Exception:
            pass
    if not dates:
        return None

    today = datetime.date.today()
    if dates[-1] != today:
        dates.append(today); prices.append(wti)

    fcast, d = [], today
    while len(fcast) < 5:
        d += datetime.timedelta(days=1)
        if d.weekday() < 5:
            fcast.append(d)
    n = len(fcast)
    f_m = [wti + (f_mid-wti)*(i+1)/n for i in range(n)]
    f_l = [wti + (f_low-wti)*(i+1)/n for i in range(n)]
    f_h = [wti + (f_high-wti)*(i+1)/n for i in range(n)]

    fig, ax = plt.subplots(figsize=(14, 6))
    fig.patch.set_facecolor('#0d1117'); ax.set_facecolor('#0d1117')
    ax.grid(color='#1e2936', linewidth=0.5, linestyle='--', alpha=0.6)
    ax.plot(dates, prices, color='#00d4ff', linewidth=2.0, label='WTI (actual)', zorder=5)
    ax.fill_between(dates, prices, min(prices)*0.97, alpha=0.10, color='#00d4ff')
    if len(prices) >= 7:
        ma7 = np.convolve(prices, np.ones(7)/7, mode='valid')
        ax.plot(dates[6:], ma7, color='#ffd700', linewidth=1.1, linestyle='--', alpha=0.75, label='7d MA')
    if len(prices) >= 20:
        ma20 = np.convolve(prices, np.ones(20)/20, mode='valid')
        ax.plot(dates[19:], ma20, color='#ff6b35', linewidth=1.1, linestyle=':', alpha=0.75, label='20d MA')
    conn_d = [today]+fcast; conn_m=[wti]+f_m; conn_l=[wti]+f_l; conn_h=[wti]+f_h
    fc = '#00ff88' if f_mid >= wti else '#ff4444'
    ax.fill_between(conn_d, conn_l, conn_h, alpha=0.15, color=fc)
    ax.plot(conn_d, conn_m, color=fc, linewidth=2.0, linestyle='--', label='1-wk forecast')
    ax.plot(conn_d, conn_l, color=fc, linewidth=0.7, linestyle=':', alpha=0.6)
    ax.plot(conn_d, conn_h, color=fc, linewidth=0.7, linestyle=':', alpha=0.6)
    ax.axvline(today, color='#555', linewidth=1, linestyle='--', alpha=0.8)
    ax.scatter([today], [wti], color='#fff', s=60, zorder=10)
    ax.annotate(' ${:.2f}'.format(wti), xy=(today, wti), color='#fff', fontsize=9, fontweight='bold', va='center')
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
    ax.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=0))
    plt.xticks(rotation=30, color='#8b9ab5', fontsize=8)
    plt.yticks(color='#8b9ab5', fontsize=9)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x,_:'${:.0f}'.format(x)))
    for sp in ax.spines.values(): sp.set_edgecolor('#1e2936')
    fig.suptitle('WTI CRUDE OIL -- 90-DAY HISTORY & 1-WEEK FORECAST',
                 color='#e8edf5', fontsize=13, fontweight='bold', fontfamily='monospace', y=0.97)
    ax.set_title('Forecast: ${:.1f}-${:.1f}/bbl  |  Model: log-normal + bootstrap ensemble'.format(f_low, f_high),
                 color='#8b9ab5', fontsize=9)
    ax.set_ylabel('Price (USD/bbl)', color='#8b9ab5', fontsize=10)
    ax.legend(loc='upper left', framealpha=0.2, facecolor='#0d1117',
              edgecolor='#1e2936', labelcolor='#c8d0de', fontsize=8)
    plt.tight_layout()
    return _b64(fig)


def chart_prob(wti, prob_table):
    horizons = list(prob_table.keys())
    # Only show bins where max probability > 0.5% to keep chart clean
    all_bins = [b for b,_ in prob_table[horizons[0]]]
    fig, axes = plt.subplots(1, len(horizons), figsize=(16, 5), sharey=True)
    fig.patch.set_facecolor('#0d1117')
    colors = ['#00d4ff','#00b4e0','#0090c0','#0070a0','#004f80']
    for idx, (h, ax) in enumerate(zip(horizons, axes)):
        probs = [p*100 for _,p in prob_table[h]]
        bars  = ax.barh(all_bins, probs, color=colors[idx], alpha=0.85, edgecolor='#0d1117')
        ax.set_facecolor('#0d1117')
        ax.set_title(h, color='#e8edf5', fontsize=11, fontweight='bold')
        ax.tick_params(colors='#8b9ab5', labelsize=7)
        ax.set_xlabel('Prob (%)', color='#8b9ab5', fontsize=8)
        for sp in ax.spines.values(): sp.set_edgecolor('#1e2936')
        for bar, p in zip(bars, probs):
            if p > 1.5:
                ax.text(bar.get_width()+0.3, bar.get_y()+bar.get_height()/2,
                        '{:.1f}%'.format(p), va='center', color='#c8d0de', fontsize=6)
    fig.suptitle('WTI PRICE PROBABILITY DISTRIBUTION BY HORIZON',
                 color='#e8edf5', fontsize=12, fontweight='bold', y=1.01)
    plt.tight_layout()
    return _b64(fig)


# -- GAS EXCEL -----------------------------------------------------------------

def generate_gas_excel(wti, returns, output_path):
    gas = wti / 42.0 + WTI_TO_GAS_MARKUP
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Gas Price Probabilities"
    hf = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    af = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    wf = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    gf = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    ws.merge_cells("A1:G1")
    ws["A1"].value = "US Retail Gas Price Probability  |  WTI: ${:.2f}/bbl  |  Gas ref: ${:.2f}/gal  |  {}".format(
        wti, gas, datetime.date.today())
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color="1F4E79")
    ws["A1"].alignment = center; ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:G2")
    ws["A2"].value = "Model: 55% log-normal + 45% bootstrap  |  Markup: WTI/42 + ${:.2f}/gal".format(WTI_TO_GAS_MARKUP)
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="7F7F7F")
    ws["A2"].alignment = center; ws.row_dimensions[2].height = 14

    for ci, h in enumerate(["Price ($/gal)"] + HORIZONS, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = hf; c.alignment = center; c.border = bdr
    ws.row_dimensions[4].height = 20

    for ri, lbl in enumerate(GAS_PRICE_BINS, 5):
        fill = af if ri % 2 == 0 else wf
        lc = ws.cell(row=ri, column=1, value=lbl)
        lc.font = Font(name="Arial", bold=True, size=10)
        lc.fill = fill; lc.alignment = left; lc.border = bdr
        for ci, h in enumerate(HORIZONS, 2):
            try:
                p1 = lognormal_probs(gas, returns, HORIZON_DAYS[h], GAS_BIN_EDGES)
                p2 = bootstrap_probs(gas, returns, HORIZON_DAYS[h], GAS_BIN_EDGES, n=2000)
                prob = ensemble(p1, p2)[ri-5]
            except Exception:
                prob = 1.0/len(GAS_PRICE_BINS)
            c = ws.cell(row=ri, column=ci, value=round(prob, 4))
            c.number_format = "0.00%"
            c.font = Font(name="Arial", size=10)
            c.fill = fill; c.alignment = center; c.border = bdr
        ws.row_dimensions[ri].height = 18

    last = 4 + len(GAS_PRICE_BINS); sr = last + 1
    ws.cell(row=sr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr, column=1).fill = gf
    ws.cell(row=sr, column=1).alignment = left
    ws.cell(row=sr, column=1).border = bdr
    for ci in range(2, 7):
        cl = get_column_letter(ci)
        c = ws.cell(row=sr, column=ci, value="=SUM({}5:{}{})".format(cl, cl, last))
        c.number_format = "0.00%"
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = gf; c.alignment = center; c.border = bdr
    ws.column_dimensions["A"].width = 16
    for col in ["B","C","D","E","F"]: ws.column_dimensions[col].width = 10
    nr = sr + 2
    ws.merge_cells("A{}:F{}".format(nr, nr))
    ws["A{}".format(nr)].value = "Note: Model-based probabilities. Gas = WTI/42 + markup. Not financial advice."
    ws["A{}".format(nr)].font = Font(name="Arial", italic=True, size=8, color="7F7F7F")
    ws["A{}".format(nr)].alignment = left
    wb.save(output_path)
    return output_path


# -- MAIN ENTRY ----------------------------------------------------------------

def run(send=print):
    send("=== OIL PRICE AGENT v2 (deterministic) ===")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts      = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = os.path.join(OUTPUT_DIR, "{}_{}".format(ts, uuid.uuid4().hex[:8]))
    os.makedirs(run_dir, exist_ok=True)
    send("  Output: {}".format(run_dir))

    raw     = _df.fetch_all(["WTI", "Brent"], send=send)
    wti_d   = raw.get("WTI",   {})
    brent_d = raw.get("Brent", {})
    wti     = wti_d.get("current", 78.0)
    brent   = brent_d.get("current") or round(wti + 2.0, 2)
    history = wti_d.get("history", [])
    returns = wti_d.get("returns", [])

    today_str = str(datetime.date.today())
    if history and history[-1]["date"] != today_str:
        history.append({"date": today_str, "price": wti})
    elif not history:
        history = [{"date": today_str, "price": wti}]

    send("  WTI ${:.2f}  Brent ${:.2f}  {} history pts".format(wti, brent, len(history)))

    send("Running models ...")
    forecast   = compute_forecast(wti, returns)
    prob_table = build_wti_prob_table(wti, returns)
    summary    = market_summary(wti, brent, history)

    send("  {} | ${:.1f}-${:.1f} | vol {:.1f}%".format(
        forecast["direction"], forecast["forecast_low"],
        forecast["forecast_high"], forecast["annualised_vol"]))

    # -- Extended Analytics ----------------------------------------------------
    send("Computing extended analytics ...")
    r_arr     = np.array(returns) if len(returns) > 5 else np.random.normal(0, 0.012, 60)
    sig_daily = float(np.std(r_arr, ddof=1)) or 0.012
    sig_daily = min(sig_daily, 0.80 / np.sqrt(252))
    from scipy.stats import norm as _norm

    # 1. CI Bands
    ci_bands = {}
    for h in HORIZONS:
        days = HORIZON_DAYS[h]
        lm = np.log(wti) - 0.5 * sig_daily**2 * days
        ls = sig_daily * np.sqrt(days)
        ci_bands[h] = {
            "mid":  round(float(np.exp(lm)), 2),
            "ci80": [round(float(np.exp(lm - 1.28*ls)), 2), round(float(np.exp(lm + 1.28*ls)), 2)],
            "ci90": [round(float(np.exp(lm - 1.645*ls)), 2), round(float(np.exp(lm + 1.645*ls)), 2)],
            "ci95": [round(float(np.exp(lm - 1.96*ls)), 2), round(float(np.exp(lm + 1.96*ls)), 2)],
        }

    # 2. Volatility heatmap
    vol_heatmap = []
    if len(history) >= 11:
        for i in range(10, len(history)):
            chunk = r_arr[max(0, i-10):i]
            if len(chunk) >= 2:
                rv = float(np.std(chunk, ddof=1)) * np.sqrt(252) * 100
                vol_heatmap.append({"date": history[i]["date"], "vol": round(rv, 2)})

    # 3. Driver analysis
    brent_spread = abs((brent or wti+2) - wti)
    seasonal_m   = datetime.date.today().month
    raw_drvs = [
        ("Supply / OPEC",       round(min(15, brent_spread * 1.5), 2)),
        ("Brent-WTI Spread",    round(min(12, brent_spread * 2.0), 2)),
        ("Seasonal Demand",     round(6.0 if seasonal_m in (6,7,8) else 3.0, 2)),
        ("USD Strength",        round(4.5, 2)),
        ("Geopolitical Risk",   round(max(2.0, forecast["annualised_vol"] / 8), 2)),
        ("Refinery Utilisation",round(5.0, 2)),
    ]
    tot = sum(v for _, v in raw_drvs) or 1
    drivers = [{"name": n, "value": v, "pct": round(v/tot*100, 1)} for n, v in raw_drvs]

    # 4. Scenario simulation
    np.random.seed(42)
    fc_dates = []
    d = datetime.date.today()
    while len(fc_dates) < 14:
        d += datetime.timedelta(days=1)
        if d.weekday() < 5:
            fc_dates.append(str(d))
    scenario_defs = {
        "Base":            {"drift":  0.000, "vol_mult": 1.0},
        "High Demand":     {"drift":  0.004, "vol_mult": 1.2},
        "Supply Disruption":{"drift": 0.008, "vol_mult": 2.0},
        "Stable Market":   {"drift":  0.001, "vol_mult": 0.6},
        "Recession":       {"drift": -0.006, "vol_mult": 1.5},
    }
    scenario_paths = {}
    for sname, sp in scenario_defs.items():
        path = [wti]
        for _ in range(14):
            path.append(round(float(path[-1] * np.exp(np.random.normal(sp["drift"], sig_daily * sp["vol_mult"]))), 2))
        scenario_paths[sname] = {"dates": fc_dates, "prices": path[1:], "final": round(path[-1], 2)}

    # 5. Regional prices (gas retail equivalent)
    gas_base = wti / 42.0 + WTI_TO_GAS_MARKUP
    regional_prices = [
        {"region": "New England",  "state": "CT", "lat": 41.6, "lon": -72.7, "price": round(gas_base * 1.18, 3), "factor": "High taxes, port logistics"},
        {"region": "Mid-Atlantic", "state": "NY", "lat": 40.7, "lon": -74.0, "price": round(gas_base * 1.13, 3), "factor": "Urban premium"},
        {"region": "Southeast",    "state": "GA", "lat": 33.7, "lon": -84.4, "price": round(gas_base * 0.96, 3), "factor": "Low state tax"},
        {"region": "Midwest",      "state": "IL", "lat": 41.8, "lon": -87.6, "price": round(gas_base * 1.04, 3), "factor": "Inland logistics"},
        {"region": "Gulf Coast",   "state": "TX", "lat": 29.7, "lon": -95.4, "price": round(gas_base * 0.90, 3), "factor": "Refinery proximity"},
        {"region": "West Coast",   "state": "CA", "lat": 34.0, "lon": -118.2,"price": round(gas_base * 1.28, 3), "factor": "CARB spec + high tax"},
        {"region": "Pacific NW",   "state": "WA", "lat": 47.6, "lon": -122.3,"price": round(gas_base * 1.17, 3), "factor": "Remote supply chain"},
        {"region": "Mountain",     "state": "CO", "lat": 39.7, "lon": -104.9,"price": round(gas_base * 1.08, 3), "factor": "Altitude distribution"},
    ]

    # 6. Profit/cost impact
    REFINE_COST = 8.0
    CRACK_MARGIN = 18.0
    retail_bbl = wti + REFINE_COST + CRACK_MARGIN
    cost_history = [
        {"date": r["date"], "cost": round(float(r["price"]) + REFINE_COST, 2),
         "retail": round(float(r["price"]) + REFINE_COST + CRACK_MARGIN, 2),
         "margin_pct": round(CRACK_MARGIN / (float(r["price"]) + REFINE_COST + CRACK_MARGIN) * 100, 2)}
        for r in history[-90:]
    ]
    profit_impact = {
        "retail": round(retail_bbl, 2), "breakeven": round(wti + REFINE_COST, 2),
        "logistics": REFINE_COST, "margin": CRACK_MARGIN,
        "margin_pct": round(CRACK_MARGIN / retail_bbl * 100, 2),
        "cost_history": cost_history,
        "scenario_retail": {s: round(scenario_paths[s]["final"] + REFINE_COST + CRACK_MARGIN, 2) for s in scenario_paths},
    }

    report_path = os.path.join(run_dir, "oil_report_{}.txt".format(ts))
    gas_path    = os.path.join(run_dir, "gas_price_{}.xlsx".format(ts))
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(summary + "\n\n--- FORECAST ---\n" + json.dumps(forecast, indent=2))
    send("  Report saved")
    try:
        generate_gas_excel(wti, returns, gas_path)
        send("  Gas Excel saved")
    except Exception as e:
        send("  [WARN] Excel skipped: {}".format(e))
        gas_path = None
    send("=== DONE ===")

    return {
        "agent":      "oil",
        "wti":        wti,
        "brent":      brent,
        "forecast":   forecast,
        "history":    history,
        "returns":    [round(float(r), 6) for r in r_arr.tolist()],
        "prob_table": {h: list(prob_table[h]) for h in HORIZONS},
        "summary":    summary,
        # Extended analytics
        "ci_bands":        ci_bands,
        "vol_heatmap":     vol_heatmap,
        "drivers":         drivers,
        "scenario_paths":  scenario_paths,
        "regional_prices": regional_prices,
        "profit_impact":   profit_impact,
        "files":      {"report": report_path, "gas_excel": gas_path},
        "run_dir":    run_dir,
    }


if __name__ == "__main__":
    result = run()
    print(result["summary"])
